#!/usr/bin/env python3
"""
Accounts Receivable Analysis Tool
Automated AR aging and collections analysis with exclusions for wire fees and tax withholdings

Usage:
    python ar_analysis.py input_file.xlsx [output_file.xlsx]

Requirements:
    - openpyxl
    - pandas
    - datetime

Author: AR Analysis Tool
Date: September 2025
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, date
import sys
import os

class ARAnalyzer:
    def __init__(self, input_file, output_file=None):
        self.input_file = input_file
        self.output_file = output_file or f"ar_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        # Configuration - customize these as needed
        self.wire_fee_threshold = 100  # Amounts <= this are considered wire fees
        self.india_withholding_docs = ['3148']  # Document numbers for India withholding tax
        self.exclusion_notes = {
            'wire_fee': 'Remaining wire fees - excluded from AR',
            'india_withholding': 'India Withholding Tax - excluded from AR'
        }
        
    def load_data(self):
        """Load and validate input data"""
        try:
            self.df = pd.read_excel(self.input_file)
            print(f"✓ Loaded {len(self.df)} records from {self.input_file}")
            
            # Expected columns (adjust based on your data structure)
            required_cols = ['Document Number', 'Name', 'Invoice Date', 'Due Date', 'Payment Date', 'Amount']
            
            # Map common column variations
            column_mapping = {
                'Maximum of Date': 'Invoice Date',
                'Maximum of Due Date/Receive By': 'Due Date', 
                'Maximum of Payment Date': 'Payment Date',
                'Sum of Amount': 'Amount'
            }
            
            # Rename columns if needed
            self.df.rename(columns=column_mapping, inplace=True)
            
            return True
        except Exception as e:
            print(f"✗ Error loading data: {e}")
            return False
    
    def calculate_days_past_due(self):
        """Calculate correct days past due"""
        def days_past_due_calc(row):
            due_date = row['Due Date']
            payment_date = row['Payment Date']
            
            if pd.notna(payment_date):
                # If payment was made, calculate based on payment date vs due date
                return max(0, (payment_date - due_date).days)
            else:
                # If no payment, calculate based on today vs due date
                today = datetime.now().date()
                due_date_only = due_date.date() if pd.notna(due_date) else today
                return max(0, (today - due_date_only).days)
        
        self.df['Days Past Due'] = self.df.apply(days_past_due_calc, axis=1)
        print("✓ Calculated days past due")
    
    def categorize_invoices(self):
        """Categorize invoices as paid, collectible AR, or excluded"""
        self.df['Category'] = 'Unknown'
        self.df['Exclusion Reason'] = ''
        
        paid_count = 0
        wire_fee_count = 0
        india_withholding_count = 0
        collectible_count = 0
        
        for idx, row in self.df.iterrows():
            doc_num = str(row['Document Number'])
            amount = row['Amount'] or 0
            payment_date = row['Payment Date']
            
            if pd.notna(payment_date):
                # Paid invoice
                self.df.at[idx, 'Category'] = 'Paid'
                paid_count += 1
            elif doc_num in self.india_withholding_docs:
                # India withholding tax
                self.df.at[idx, 'Category'] = 'Excluded'
                self.df.at[idx, 'Exclusion Reason'] = self.exclusion_notes['india_withholding']
                india_withholding_count += 1
            elif amount <= self.wire_fee_threshold:
                # Wire fee
                self.df.at[idx, 'Category'] = 'Excluded'
                self.df.at[idx, 'Exclusion Reason'] = self.exclusion_notes['wire_fee']
                wire_fee_count += 1
            else:
                # Collectible AR
                self.df.at[idx, 'Category'] = 'Collectible AR'
                collectible_count += 1
        
        print(f"✓ Categorized invoices:")
        print(f"  - Paid: {paid_count}")
        print(f"  - Collectible AR: {collectible_count}")
        print(f"  - Wire fees excluded: {wire_fee_count}")
        print(f"  - India withholding excluded: {india_withholding_count}")
    
    def calculate_ar_metrics(self):
        """Calculate key AR metrics"""
        # Separate data
        paid_invoices = self.df[self.df['Category'] == 'Paid']
        collectible_ar = self.df[self.df['Category'] == 'Collectible AR']
        excluded = self.df[self.df['Category'] == 'Excluded']
        
        # Key metrics
        self.metrics = {
            'total_payments': paid_invoices['Amount'].sum(),
            'collectible_ar': collectible_ar['Amount'].sum(),
            'excluded_total': excluded['Amount'].sum(),
            'wire_fees': excluded[excluded['Exclusion Reason'].str.contains('wire fees', na=False)]['Amount'].sum(),
            'india_withholding': excluded[excluded['Exclusion Reason'].str.contains('India', na=False)]['Amount'].sum(),
            'paid_count': len(paid_invoices),
            'collectible_count': len(collectible_ar),
            'excluded_count': len(excluded),
            'collection_rate': len(paid_invoices) / (len(paid_invoices) + len(collectible_ar)) * 100,
            'oldest_days': collectible_ar['Days Past Due'].max() if len(collectible_ar) > 0 else 0
        }
        
        # AR Aging
        def categorize_aging(days_past_due):
            if days_past_due == 0:
                return 'On Time'
            elif days_past_due <= 30:
                return '1-30 Days Past Due'
            elif days_past_due <= 60:
                return '31-60 Days Past Due'
            elif days_past_due <= 90:
                return '61-90 Days Past Due'
            else:
                return 'Over 90 Days Past Due'
        
        collectible_ar['Aging Category'] = collectible_ar['Days Past Due'].apply(categorize_aging)
        
        self.aging_summary = collectible_ar.groupby('Aging Category').agg({
            'Amount': ['sum', 'count']
        }).round(2)
        self.aging_summary.columns = ['Total Amount', 'Invoice Count']
        self.aging_summary = self.aging_summary.reset_index()
        
        # Order properly
        aging_order = ['On Time', '1-30 Days Past Due', '31-60 Days Past Due', '61-90 Days Past Due', 'Over 90 Days Past Due']
        self.aging_summary['Order'] = self.aging_summary['Aging Category'].map({cat: i for i, cat in enumerate(aging_order)})
        self.aging_summary = self.aging_summary.sort_values('Order').drop('Order', axis=1)
        
        print(f"✓ Calculated AR metrics:")
        print(f"  - Collectible AR: ${self.metrics['collectible_ar']:,.2f}")
        print(f"  - Collection Rate: {self.metrics['collection_rate']:.1f}%")
    
    def create_excel_report(self):
        """Create comprehensive Excel report"""
        # Sort data by days past due (descending)
        df_sorted = self.df.sort_values('Days Past Due', ascending=False)
        
        # Create workbook
        wb = openpyxl.Workbook()
        
        # Executive Summary Sheet
        self._create_executive_summary(wb)
        
        # Invoice Data Sheet  
        self._create_invoice_data_sheet(wb, df_sorted)
        
        # Collections Analysis Sheet
        self._create_collections_analysis_sheet(wb)
        
        # Save workbook
        wb.save(self.output_file)
        print(f"✓ Created Excel report: {self.output_file}")
    
    def _create_executive_summary(self, wb):
        """Create executive summary dashboard"""
        sheet = wb.create_sheet('Executive Summary', 0)
        
        # Title
        sheet['A1'] = 'ACCOUNTS RECEIVABLE EXECUTIVE SUMMARY'
        sheet['A1'].font = Font(bold=True, size=18)
        sheet.merge_cells('A1:F1')
        sheet['A1'].alignment = Alignment(horizontal='center')
        
        sheet['A2'] = f'As of {datetime.now().strftime("%B %d, %Y")}'
        sheet['A2'].font = Font(size=12, color='666666')
        sheet.merge_cells('A2:F2')
        sheet['A2'].alignment = Alignment(horizontal='center')
        
        # Key Metrics
        sheet['A4'] = 'KEY METRICS'
        sheet['A4'].font = Font(bold=True, size=14)
        sheet['A4'].fill = PatternFill('solid', start_color='D3D3D3')
        
        metrics = [
            ['Total Payments Received:', f"${self.metrics['total_payments']:,.2f}"],
            ['Collectible Outstanding AR:', f"${self.metrics['collectible_ar']:,.2f}"],
            ['Excluded Items:', f"${self.metrics['excluded_total']:,.2f}"],
            ['  - India Withholding Tax:', f"${self.metrics['india_withholding']:,.2f}"],
            ['  - Wire Fees:', f"${self.metrics['wire_fees']:,.2f}"],
            ['Collectible Outstanding Invoices:', f"{self.metrics['collectible_count']:,}"],
            ['Collection Rate (by count):', f"{self.metrics['collection_rate']:.1f}%"]
        ]
        
        for i, (metric, value) in enumerate(metrics, 5):
            sheet.cell(row=i, column=1, value=metric).font = Font(bold=True)
            sheet.cell(row=i, column=2, value=value)
        
        # Aging Summary
        sheet['A13'] = 'AGING SUMMARY'
        sheet['A13'].font = Font(bold=True, size=14)
        sheet['A13'].fill = PatternFill('solid', start_color='D3D3D3')
        
        headers = ['Category', 'Amount', 'Percentage', 'Invoice Count']
        for i, header in enumerate(headers, 1):
            cell = sheet.cell(row=14, column=i, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill('solid', start_color='E6E6FA')
        
        for row_idx, (_, row) in enumerate(self.aging_summary.iterrows(), 15):
            sheet.cell(row=row_idx, column=1, value=row['Aging Category'])
            sheet.cell(row=row_idx, column=2, value=f"${row['Total Amount']:,.2f}")
            pct = (row['Total Amount'] / self.metrics['collectible_ar']) * 100 if self.metrics['collectible_ar'] > 0 else 0
            sheet.cell(row=row_idx, column=3, value=f"{pct:.1f}%")
            sheet.cell(row=row_idx, column=4, value=f"{row['Invoice Count']:,}")
        
        # Key Findings
        over_90_amount = self.aging_summary[self.aging_summary['Aging Category'] == 'Over 90 Days Past Due']['Total Amount'].iloc[0] if not self.aging_summary[self.aging_summary['Aging Category'] == 'Over 90 Days Past Due'].empty else 0
        over_90_count = self.aging_summary[self.aging_summary['Aging Category'] == 'Over 90 Days Past Due']['Invoice Count'].iloc[0] if not self.aging_summary[self.aging_summary['Aging Category'] == 'Over 90 Days Past Due'].empty else 0
        
        findings = [
            f"• Collection Success: {self.metrics['collection_rate']:.1f}% of collectible invoices have been collected",
            f"• Outstanding AR: ${self.metrics['collectible_ar']:,.0f} in truly collectible receivables ({self.metrics['collectible_count']} invoices)",
            f"• High Risk: ${over_90_amount:,.0f} over 90 days past due across {over_90_count} invoices",
            f"• Excluded Items: ${self.metrics['excluded_total']:,.0f} properly categorized as non-collectible"
        ]
        
        sheet['A21'] = 'KEY FINDINGS & RECOMMENDATIONS'
        sheet['A21'].font = Font(bold=True, size=14)
        sheet['A21'].fill = PatternFill('solid', start_color='D3D3D3')
        
        for i, finding in enumerate(findings, 22):
            sheet.cell(row=i, column=1, value=finding)
            sheet.merge_cells(f'A{i}:F{i}')
        
        # Recommended Actions
        sheet['A27'] = 'RECOMMENDED ACTIONS'
        sheet['A27'].font = Font(bold=True, size=14)
        sheet['A27'].fill = PatternFill('solid', start_color='FFB6C1')
        
        actions = [
            f'1. FOCUS: Target {over_90_count} invoices over 90 days past due (${over_90_amount:,.0f})',
            '2. PRIORITY: Large invoices in 31-60 day bucket need immediate attention',
            '3. INVESTIGATE: Why no current AR - all outstanding invoices are past due',
            f'4. ESCALATE: Oldest invoice ({self.metrics["oldest_days"]} days) requires legal action',
            f'5. PROCESS: Review collection process - {100-self.metrics["collection_rate"]:.0f}% of invoices remain uncollected'
        ]
        
        for i, action in enumerate(actions, 28):
            sheet.cell(row=i, column=1, value=action)
            sheet.merge_cells(f'A{i}:F{i}')
    
    def _create_invoice_data_sheet(self, wb, df_sorted):
        """Create detailed invoice data sheet"""
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        sheet = wb.create_sheet('Invoice Data')
        
        # Headers
        headers = ['Document Number', 'Name', 'Invoice Date', 'Due Date', 'Payment Date', 'Amount', 'Days Past Due', 'Notes']
        for i, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=i, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill('solid', start_color='E6E6FA')
        
        # Data
        for row_idx, (_, row) in enumerate(df_sorted.iterrows(), 2):
            sheet.cell(row=row_idx, column=1, value=row['Document Number'])
            sheet.cell(row=row_idx, column=2, value=row['Name'])
            
            # Dates with proper formatting
            for col, date_field in [(3, 'Invoice Date'), (4, 'Due Date'), (5, 'Payment Date')]:
                if pd.notna(row[date_field]):
                    cell = sheet.cell(row=row_idx, column=col, value=row[date_field])
                    cell.number_format = 'MM/DD/YYYY'
            
            # Amount with currency formatting
            amount_cell = sheet.cell(row=row_idx, column=6, value=row['Amount'])
            amount_cell.number_format = '"$"#,##0.00'
            
            # Days past due
            sheet.cell(row=row_idx, column=7, value=row['Days Past Due'])
            
            # Notes
            if row['Category'] == 'Excluded':
                note_cell = sheet.cell(row=row_idx, column=8, value=row['Exclusion Reason'])
                note_cell.font = Font(italic=True, color='666666')
        
        # Column widths
        column_widths = {'A': 15, 'B': 40, 'C': 12, 'D': 12, 'E': 12, 'F': 15, 'G': 12, 'H': 35}
        for col, width in column_widths.items():
            sheet.column_dimensions[col].width = width
    
    def _create_collections_analysis_sheet(self, wb):
        """Create collections analysis sheet"""
        sheet = wb.create_sheet('Collections Analysis')
        
        sheet['A1'] = 'Outstanding AR Analysis by Aging'
        sheet['A1'].font = Font(bold=True, size=16)
        sheet.merge_cells('A1:D1')
        
        # Headers
        headers = ['Aging Category', 'Total Amount', 'Invoice Count', 'Percentage']
        for i, header in enumerate(headers, 1):
            cell = sheet.cell(row=3, column=i, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill('solid', start_color='E6E6FA')
        
        # Data
        for row_idx, (_, row) in enumerate(self.aging_summary.iterrows(), 4):
            sheet.cell(row=row_idx, column=1, value=row['Aging Category'])
            
            amount_cell = sheet.cell(row=row_idx, column=2, value=row['Total Amount'])
            amount_cell.number_format = '"$"#,##0.00'
            
            sheet.cell(row=row_idx, column=3, value=row['Invoice Count'])
            
            pct = (row['Total Amount'] / self.metrics['collectible_ar']) * 100 if self.metrics['collectible_ar'] > 0 else 0
            pct_cell = sheet.cell(row=row_idx, column=4, value=pct/100)
            pct_cell.number_format = '0.0%'
        
        # Total row
        total_row = len(self.aging_summary) + 4
        sheet.cell(row=total_row, column=1, value='TOTAL').font = Font(bold=True)
        
        total_amount_cell = sheet.cell(row=total_row, column=2, value=f'=SUM(B4:B{total_row-1})')
        total_amount_cell.font = Font(bold=True)
        total_amount_cell.number_format = '"$"#,##0.00'
        
        total_count_cell = sheet.cell(row=total_row, column=3, value=f'=SUM(C4:C{total_row-1})')
        total_count_cell.font = Font(bold=True)
        
        sheet.cell(row=total_row, column=4, value='100.0%').font = Font(bold=True)
        
        # Column widths
        column_widths = {'A': 22, 'B': 18, 'C': 15, 'D': 12}
        for col, width in column_widths.items():
            sheet.column_dimensions[col].width = width
    
    def run_analysis(self):
        """Run complete AR analysis"""
        print("🚀 Starting AR Analysis...")
        
        if not self.load_data():
            return False
        
        self.calculate_days_past_due()
        self.categorize_invoices()
        self.calculate_ar_metrics()
        self.create_excel_report()
        
        print(f"\n✅ Analysis Complete!")
        print(f"📊 Report saved as: {self.output_file}")
        print(f"💰 Collectible AR: ${self.metrics['collectible_ar']:,.2f}")
        print(f"📈 Collection Rate: {self.metrics['collection_rate']:.1f}%")
        
        return True

def main():
    """Main function for command line usage"""
    if len(sys.argv) < 2:
        print("Usage: python ar_analysis.py input_file.xlsx [output_file.xlsx]")
        print("Example: python ar_analysis.py invoices.xlsx ar_report.xlsx")
        return
    
    input_file = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) > 2 else None
    
    if not os.path.exists(input_file):
        print(f"Error: Input file '{input_file}' not found")
        return
    
    analyzer = ARAnalyzer(input_file, output_file)
    analyzer.run_analysis()

if __name__ == "__main__":
    main()
